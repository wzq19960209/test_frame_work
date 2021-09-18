import openpyxl
import yaml
from test_framework.setting import DIR_NAME
'''
需求：打开data.xlsx文件，提取测试用例中的数据，最后以列表嵌套列表的形式展示出来
'''

all_list=[]
class ReadeExcel():
    def get_excel_list(self):
        wb = openpyxl.load_workbook(DIR_NAME+'/data/data.xlsx')
        ws = wb['测试用例']
        for rows in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1,max_col=ws.max_column):    # iter_rows指的是行，后面的括号里面代表了这个范围的大小    #生成器可以理解为能用for ... in去循环遍历的一个迭代器
            cell_value_list = []   #定义列表的位置很重要，他会影响这个列表循环遍历的结果
            for cell in rows:
                cell_value_list.append(cell.value)
            all_list.append(cell_value_list)
        return all_list

    def get_yaml(self,filename,key):
        with open(DIR_NAME+'/data/%s.yml'%filename,'r',encoding='utf-8')  as f:
            yaml_data=yaml.safe_load(f)
            yaml_list=[]
            yaml_cases = yaml_data.get(key)  #取到test_login所对应的值，解剖第一层
            for i in yaml_cases.values():   #将yaml_cases的值循环遍历，解剖第二层，得到我们需要的测试用例数据
                yaml_list.append(i)  #将数据添加到列表中，并返回该列表
        return yaml_list


# 工具，提取数据
class ReadeExcel1():
    def get_excel(self,header):
        # excel文档
        wb = openpyxl.load_workbook(DIR_NAME+'/data/data.xlsx')
        #shell页
        ws = wb['测试用例']
        # 列表容器
        all_case = []
        #列数   是header里面元素的个数是一致的
        #父循环控制循环的行数  ,这一步需要思考:这个循环为的是构造数据还是确定循环次数
        for row in range(2,ws.max_row+1):
            #用子循环控制列数  1.主要循环次数,就可以确定现在这次循环是处于第几列     2.也要数据,key,构造字典数据
            #构造空字典
            dict={}
            #自定义控制次数
            col=1
            for i in header:
                #构造字典
                dict[i] = ws.cell(row,col).value
                col+=1
            all_case.append(dict)
        return all_case


if __name__ == '__main__':
    # li=ReadeExcel().get_excel_list()
    # print(li)
    # yaml_data=ReadeExcel().get_yaml()
    # print(yaml_data)
    print(ReadeExcel().get_yaml('login_data','test_login'))